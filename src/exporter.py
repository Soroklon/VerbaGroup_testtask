from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("url", "Ссылка на товар"),
    ("article", "Артикул"),
    ("name", "Название"),
    ("price", "Цена, ₽"),
    ("description", "Описание"),
    ("images", "Изображения"),
    ("characteristics", "Характеристики"),
    ("seller_name", "Продавец"),
    ("seller_url", "Ссылка на продавца"),
    ("sizes", "Размеры"),
    ("stock", "Остаток"),
    ("rating", "Рейтинг"),
    ("reviews", "Отзывы"),
]

COL_WIDTHS = [50, 12, 40, 10, 60, 70, 55, 30, 45, 30, 10, 10, 10]


def save_xlsx(products: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")

    for col, (_, header) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row_idx, product in enumerate(products, 2):
        for col, (key, _) in enumerate(COLUMNS, 1):
            value = product.get(key, "")
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Сохранено: {path} ({len(products)} товаров)")
